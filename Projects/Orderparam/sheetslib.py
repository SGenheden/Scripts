# Author: Samuel Genheden samuel.genheden@gmail.com

import openpyxl as xl

def _extract_residues(sheet, rowstart=1, sysoffset=0):

    n = rowstart
    reserial = 0
    while True:
        if sheet.cell(column=1+sysoffset,row=n+1).value is None or \
            len(sheet.cell(column=1+sysoffset,row=n+1).value) == 0 :
            break
        n += 1

    lst = [str(sheet.cell(column=1+sysoffset,row=i).value)+str(sheet.cell(column=2+sysoffset,row=i).value)
                +"-%d"%i
                for i in range(rowstart+1,n+1)]
    return lst

def _extract_data(sheet, residues, sysoffset=0, coloffset=0):

    col = 3
    nres = len(residues)
    labels =  []
    data = []
    while True:
        if sheet.cell(column=col+sysoffset,row=1).value is None or \
            len(sheet.cell(column=col+sysoffset,row=1).value) == 0 :
            break
        labels.append(str(sheet.cell(column=col+sysoffset,row=1).value))
        s2 = {}
        for i in range(2,nres+2):
            try :
                val = float(sheet.cell(column=col+sysoffset+coloffset,row=i).value)
            except:
                pass
            else:
                s2[residues[i-2]] = val
        data.append(s2)
        col += 2
    return labels,data

def open_book(filename):

    wb = None
    try :
        wb = xl.load_workbook(filename = filename)
    except :
        raise Exception("Could not open XLSX file")
    return wb

def extract_sheet(wb, sheetname, sysoffset):

    ws = None
    try :
        ws = wb[sheetname]
    except :
        raise Exception("Could not find sheet")

    res = _extract_residues(ws, sysoffset=sysoffset)
    labels, data = _extract_data(ws, res, sysoffset=sysoffset)
    labels, errors = _extract_data(ws, res, sysoffset=sysoffset, coloffset=1)
    return res, labels, data, errors
