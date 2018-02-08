# Author: Samuel Genheden samuel.genheden@gmail.com

"""
Program to collect TI results and put them in an Excel sheet

Examples:
collect_results.py -db MNSol_alldata.txt -solvent hexanol -solutes hexanolwater.txt
                    -xls out.xlsx -sheet hexanolwater
"""

import argparse
import os

import openpyxl as xl
import numpy as np
import scipy.stats as stats

from sgenlib import parsing

import dblib

TOKJMOL = 4.184

def _integrate(filename):
    """
    Read a PMF from an Lammps output file and integrate it
    """
    lines = []
    with open(filename,"r") as f :
        lines = f.readlines()
    lines.append("0.0 1.0 0.0")
    data = np.array([line.strip().split() for line in lines[2:]],dtype=float)
    lam = data[:,1]
    grad = data[:,2]
    # Linear interpolation to lambda=1 from the last two lambda values simulated
    grad[-1] = np.polyfit(lam[-3:-1],grad[-3:-1],1).sum()
    return np.trapz(grad,x=lam)

def _dobar(filename):
    """
    Read an oputput file from g_bar and return the dG in kcal/mol
    """
    data = parsing.parse2ndarray(filename)
    return 1.987*0.3*data[:,1].sum()

def _lammps_filename(solvent, filehandle) :
    return "out.dPotEngSS_%s_%s"%(solvent ,filehandle)

def _gmx_filename(solvent, filehandle) :
    return "%s_bar.xvg"%filehandle

def _fill_solutedata(db,solvent,simsolvent,solutes,outdir,ws, coloffset,
                        filename_fnc, dg_func):
    """
    Fill in solute name, calculate and experimental as well as
    absolute difference for all solutes. Also fills in statistical data.
    """
    ws.cell(row=1,column=1+coloffset).value = solvent.capitalize()
    _add_colhead(ws,coloffset)

    data = []
    dbentries = []
    for i,entry in enumerate(db.itersolutelist(solvent,solutes),3):
        filename = os.path.join(outdir,
                    filename_fnc(simsolvent, entry.FileHandle))
        dglist = [dg_func(filename)]
        for repeat in range(2,5):
            filename = os.path.join(outdir,"R%d"%repeat,
                        filename_fnc(simsolvent, entry.FileHandle))
            try:
                dg_repeat = dg_func(filename)
            except:
                nomore = True
            else:
                dglist.append(dg_repeat)
                nomore = False
            if nomore: break
        if len(dglist) == 1:
            dg = dglist[0]
            std = 0.0
        else:
            dg = np.asarray(dglist).mean()
            std = np.asarray(dglist).std()/np.sqrt(len(dglist))
        ws.cell(row=i,column=1+coloffset).value = entry.SoluteName
        ws.cell(row=i,column=2+coloffset).value = -dg*TOKJMOL
        ws.cell(row=i,column=3+coloffset).value = std*TOKJMOL
        ws.cell(row=i,column=4+coloffset).value = float(entry.DeltaGsolv)*TOKJMOL
        calcol = ws.cell(row=i,column=2+coloffset).column
        stdcol = ws.cell(row=i,column=3+coloffset).column
        expcol = ws.cell(row=i,column=4+coloffset).column
        ws.cell(row=i,column=5+coloffset).value = "=ABS(%s%i-%s%i)"%(calcol,i,expcol,i)
        data.append([-dg,float(entry.DeltaGsolv)])
        dbentries.append(entry)

    data = np.asarray(data)
    _fill_stats(ws,coloffset,data.shape[0],calcol,expcol,data)

    return data,dbentries,calcol,stdcol,expcol

def _add_colhead(ws,coloffset):
    """
    Fill in column heads
    """
    ws.cell(row=2,column=1+coloffset).value = "Solute"
    ws.cell(row=2,column=2+coloffset).value = "Calc."
    ws.cell(row=2,column=3+coloffset).value = "Std"
    ws.cell(row=2,column=4+coloffset).value = "Exp."
    ws.cell(row=2,column=5+coloffset).value = "Diff."

def _fill_stats(ws,coloffset,nentries,calcol,expcol,data,tau=None):
    """
    Fill in statistical analysis: MAD, r2, tau, r, slope and intercept
    """
    diffcol = ws.cell(row=3,column=5+coloffset).column
    ws.cell(row=3,column=6+coloffset).value = "MAD"
    ws.cell(row=3,column=7+coloffset).value = "=AVERAGE(%s3:%s%d)"%(diffcol,diffcol,nentries+2)

    ws.cell(row=4,column=6+coloffset).value = "r2"
    ws.cell(row=4,column=7+coloffset).value = "=CORREL(%s3:%s%d,%s3:%s%d)^2"% \
        (calcol,calcol,nentries+2,expcol,expcol,nentries+2)

    if tau is None:
        tau = stats.kendalltau(data[:,0],data[:,1])
    ws.cell(row=5,column=6+coloffset).value = "tau"
    ws.cell(row=5,column=7+coloffset).value = tau[0]
    ws.cell(row=5,column=8+coloffset).value = tau[1]

    ws.cell(row=6,column=6+coloffset).value = "r"
    ws.cell(row=6,column=7+coloffset).value = "=CORREL(%s3:%s%d,%s3:%s%d)"% \
        (calcol,calcol,nentries+2,expcol,expcol,nentries+2)
    ws.cell(row=6,column=8+coloffset).value = stats.pearsonr(data[:,0],data[:,1])[1]

    ws.cell(row=7,column=6+coloffset).value = "Slope"
    ws.cell(row=7,column=7+coloffset).value = "=SLOPE(%s3:%s%d,%s3:%s%d)^2"% \
        (expcol,expcol,nentries+2,calcol,calcol,nentries+2)

    ws.cell(row=8,column=6+coloffset).value = "Intercept"
    ws.cell(row=8,column=7+coloffset).value = "=INTERCEPT(%s3:%s%d,%s3:%s%d)^2"% \
        (expcol,expcol,nentries+2,calcol,calcol,nentries+2)

if __name__ == '__main__':

    argparser = argparse.ArgumentParser(description="Script to collect output, do TI and put in an Excel sheet")
    argparser.add_argument('-db', '--db', help="the molecule database")
    argparser.add_argument('-solvent', '--solvent', help="the solvent")
    argparser.add_argument('-simsolvent', '--simsolvent', help="the simulated solvent")
    argparser.add_argument('-solutes','--solutes',help="the list of solutes")
    argparser.add_argument('-outdir','--outdir',help="the directory with output files",default=".")
    argparser.add_argument('-watdir','--watdir',help="the directory with water output directory",default="../Water")
    argparser.add_argument('-xls','--xls',help="the filename of the XLS file")
    argparser.add_argument('-sheet','--sheet',help="the sheet in the XLS file")
    argparser.add_argument('-md','--md',choices=["lammps", "gmx"],help="the MD code used",default="lammps")
    argparser.add_argument('--calcdiff',action="store_true",help="turn on difference calculations",default=False)
    args = argparser.parse_args()

    if args.md == "lammps" :
        dg_func = _integrate
        filename_func = _lammps_filename
    else :
        dg_func = _dobar
        filename_func = _gmx_filename

    db = dblib.SolvDb(filename=args.db,type="abs",filehandle="^0")
    solutes = [s.strip() for s in open(args.solutes,'r').readlines()]

    if args.simsolvent is None :
        args.simsolvent = args.solvent

    wb = None
    try :
        wb = xl.load_workbook(filename = args.xls)
    except :
        print "Could not open the XLS file. Will create one from scratch"
        wb = xl.Workbook()

    ws = None
    try :
        ws = wb['raw_'+args.sheet]
    except :
        ws = wb.create_sheet(title='raw_'+args.sheet)

    data,dbentries,calcol,stdcol,expcol = \
        _fill_solutedata(db,args.solvent,args.simsolvent,solutes,
                            args.outdir,ws,0, filename_func, dg_func)

    if args.simsolvent != "water" or args.calcdiff :
        woutdir = args.watdir
        wdata,wdbentries,wcalcol,wstdcol,wexpcol = \
            _fill_solutedata(db, 'water', 'water', solutes, woutdir, ws, 10,
                                filename_func, dg_func)


        ws.cell(row=1,column=21).value = "Transfer"
        _add_colhead(ws,20)

        tdata = []
        if args.calcdiff :
            for i,(solute_data,wsolute_data,entry) in enumerate(zip(data,wdata,dbentries),3):
                ws.cell(row=i,column=21).value = entry.SoluteName
                ws.cell(row=i,column=22).value = "=%s%d"%(calcol,i)
                ws.cell(row=i,column=23).value = "=%s%d"%(stdcol,i)
                ws.cell(row=i,column=24).value = "=%s%d"%(wcalcol,i)
                tcalcol = ws.cell(row=i,column=22).column
                texpcol = ws.cell(row=i,column=24).column
                ws.cell(row=i,column=25).value = "=ABS(%s%i-%s%i)"%(tcalcol,i,texpcol,i)
                tdata.append([solute_data[0] ,  wsolute_data[0]])

        else :
            logpfac = "/(2.3*8.314/1000*298)"
            for i,(solute_data,wsolute_data,entry) in enumerate(zip(data,wdata,dbentries),3):
                ws.cell(row=i,column=21).value = entry.SoluteName
                ws.cell(row=i,column=22).value = "=(%s%d-%s%d)%s"%(wcalcol,i,calcol,i,logpfac)
                ws.cell(row=i,column=23).value = "=SQRT(%s%d^2+%s%d^2)%s"%(stdcol,i,wstdcol,i,logpfac)
                ws.cell(row=i,column=24).value = "=(%s%d-%s%d)%s"%(wexpcol,i,expcol,i,logpfac)
                tcalcol = ws.cell(row=i,column=22).column
                texpcol = ws.cell(row=i,column=24).column
                ws.cell(row=i,column=25).value = "=ABS(%s%i-%s%i)"%(tcalcol,i,texpcol,i)
                tdata.append([solute_data[0] - wsolute_data[0],
                                solute_data[1] - wsolute_data[1]])

        tdata = np.asarray(tdata)
        tdata_sign = np.sign(tdata)
        accur = ((tdata_sign[:,0]*tdata_sign[:,1])==1).sum()/float(tdata.shape[0])
        tau = [(tdata_sign[:,0]*tdata_sign[:,1]).mean(),accur]
        _fill_stats(ws,20,data.shape[0],tcalcol,texpcol,tdata,tau=tau)

    wb.save(args.xls)
