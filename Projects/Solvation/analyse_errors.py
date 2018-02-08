# Author: Samuel Genheden samuel.genheden@gmail.com

"""
Program to analyse the errors of the predictions compared to experiments.
Work with an Excel sheet created by collect_results.py
Create box plots of the error distribution and perform BEDROC and p-value
analysis of individual chemical groups
"""

import argparse
import os
from collections import namedtuple

import openpyxl as xl
import numpy as np
import matplotlib.pylab as plt
import croc
import scipy.stats as stats

import dblib

def _extract_nsolutes(ws):
    # G3 contains the MAD formula
    range_str = ws['G3'].value[9:-1]
    first,last = map(lambda s:int(s[1:]),range_str.split(':'))
    return last - first + 1

def _extract_errors(ws):
    nsolutes = _extract_nsolutes(ws)
    return np.asarray([np.abs(ws.cell(row=i,column=2).value -
                            ws.cell(row=i,column=4).value)
                        for i in range(3,nsolutes+3)])

def _extract_signederrors(ws):
    nsolutes = _extract_nsolutes(ws)
    return np.asarray([ws.cell(row=i,column=2).value -
                            ws.cell(row=i,column=4).value
                        for i in range(3,nsolutes+3)])

def _extract_errors_transf(ws):
    nsolutes = _extract_nsolutes(ws)
    return np.asarray([np.abs((ws.cell(row=i,column=2).value -
                                ws.cell(row=i,column=12).value)/(2.3*8.314/1000*298) -
                            (ws.cell(row=i,column=4).value-
                                ws.cell(row=i,column=14).value)/(2.3*8.314/1000*298))
                        for i in range(3,nsolutes+3)])

def _extract_solutes(ws):
    nsolutes = _extract_nsolutes(ws)
    return np.asarray([ws.cell(row=i,column=1).value for i in range(3,nsolutes+3)])

def _bedroc_analytic(npos,ntot):
    npos = npos / float(ntot)
    nneg = 1.0 - npos
    num1 = np.exp(npos)-nneg
    denom1 = np.exp(npos)-1
    denom2 = 1 - np.exp(-nneg)
    return (num1/denom1)-(nneg / denom2)

def _bootstrap_bedroc(data,nboots=50):
    bedrocs = np.zeros(nboots)
    for i in range(nboots):
        while True:
            idx = np.random.randint(0,data.shape[0]-1,data.shape[0])
            if np.sum(data[idx,1]) > 0 and np.sum(data[idx,1]) != data.shape[0]: break
        scoreddata = croc.ScoredData(data[idx,:])
        bedrocs[i] = croc.BEDROC(scoreddata,1.0)['BEDROC']
    return bedrocs.mean(),bedrocs.std()


if __name__ == '__main__':

    argparser = argparse.ArgumentParser(description="Analyse the errors compared to experiments")
    argparser.add_argument('-db', '--db', help="the molecule database")
    argparser.add_argument('-soldir','--soldir',help="the directory with all solute files")
    argparser.add_argument('-xls','--xls',help="the filename of the XLS file")
    argparser.add_argument('-solvents','--solvents',nargs="+",help="the solvents in the XLS file to analyse")
    argparser.add_argument('--noboxplots',action="store_true",default=False,help="turn off creation of boxplots")
    argparser.add_argument('-postfix','--postfix',help="the sheet name postfix",default="")
    argparser.add_argument('-groupname','--groupname',help="the sheet name postfix",default=".groups")
    argparser.add_argument('-supergroups','--supergroups',help="the super groups")
    args = argparser.parse_args()

    db = dblib.SolvDb(filename=args.db,type="abs",filehandle="^0")

    if args.supergroups is not None :
        with open(args.supergroups, "r") as f :
            supergroups = [line.strip() for line in f.readlines()]
        print supergroups
    else :
        supergroups = dblib.supergroups

    wb = None
    try :
        wb = xl.load_workbook(filename = args.xls)
    except :
        print "Could not open the XLS file. Exit"
        quit()

    sheets = []
    for solvent in args.solvents:
        try :
            postfix = args.postfix if solvent == 'water' else 'water'+args.postfix
            ws = wb['raw_'+solvent+postfix]
        except :
            raise Exception("Unable to find sheet for solvent%s"%solvent)
        else:
            sheets.append(ws)

    errorlist = [_extract_errors(sheet) for sheet in sheets]
    if not args.noboxplots :
        fig = plt.figure(1)
        fig.gca().boxplot(errorlist,notch=False)
        fig.gca().set_xticklabels(args.solvents)
        fig.savefig("errordist_e09.png",format="png",dpi=300)

        errorlist_transfer = [_extract_errors_transf(sheet)
            for solvent,sheet in zip(args.solvents,sheets) if solvent != 'water']
        fig = plt.figure(2)
        fig.gca().boxplot(errorlist_transfer,notch=False)
        fig.gca().set_xticklabels([solvent for solvent in args.solvents if solvent != 'water' ])
        fig.savefig("errordist_transf_e09.png",format="png",dpi=300)

    # Print out outliers
    for solvent,sheet,errors in zip(args.solvents,sheets,errorlist):
        print "\nOutliers for:\n%s\t"%solvent,
        perc90 = np.percentile(errors,90)
        idx = errors >= perc90
        print ", ".join(["%s (%.1f)"%(solute,error)
            for solute,error in zip(_extract_solutes(sheet)[idx],errors[idx])])


    # Do analysis of chemical groups
    print ""
    statslist = {group:[] for group in supergroups}
    GroupStatData = \
        namedtuple("GroupStatData",["value","mean","std","analytic","pvalue","mse","n"])
    for solvent,sheet,errors in zip(args.solvents,sheets,errorlist):
        solutes = _extract_solutes(sheet)
        signederrors = _extract_signederrors(sheet)
        entries = [entry for entry in db.itersolutelist(solvent if solvent != 'waterpol' else 'water',solutes)]
        ntrue = sum([entry.SoluteName==solute for entry,solute in zip(entries,solutes)])
        if ntrue != len(solutes) :
            raise Exception("Internal error!")
        solgroups = []
        for entry in entries:
            with open(os.path.join(args.soldir,"%s%s"%(entry.FileHandle, args.groupname)),"r") as f :
                groups = [line.strip().split(":")[0] for line in f.readlines()]
            if args.supergroups is  None:
                groups = dblib.transform_groups(groups,entry)
            solgroups.append(groups)
        for group in supergroups:
            scoreddata = croc.ScoredData()
            bootdata = np.zeros((errors.shape[0],2))
            for i,(error,solgroup) in enumerate(zip(errors,solgroups)):
                if group in solgroup :
                    scoreddata.add(error,1)
                    bootdata[i,:] = [error,1]
                else:
                    scoreddata.add(error,0)
                    bootdata[i,:] = [error,0]
            if np.sum(bootdata[:,1]) >= 5 :
                bedroc0 = croc.BEDROC(scoreddata,1.0)['BEDROC']
                bedroc_anal = _bedroc_analytic(np.sum(bootdata[:,1]),errors.shape[0])
                bedroc_mean,bedroc_std = _bootstrap_bedroc(bootdata)
                posdata = bootdata[bootdata[:,1]==1,0]
                signedposdata = signederrors[bootdata[:,1]==1]
                t,pval = stats.ttest_ind(posdata,errors,equal_var=False)
                statslist[group].append(GroupStatData(bedroc0,bedroc_mean,
                                                bedroc_std,bedroc_anal,pval,
                                                signedposdata.mean(),posdata.shape[0]))
            else:
                statslist[group].append(None)

    # Finally put all the group analysis in a sheet in the Excel document
    try :
        ws = wb['groupanal']
    except :
        ws = wb.create_sheet(title='groupanal')
    rowoffset = len(supergroups)+2

    ws['A1'] = 'Group'
    ws.cell(row=rowoffset+1,column=1).value = 'Analytical BEDROC'
    ws.cell(row=rowoffset*2+1,column=1).value = 'p-values of t-test'
    ws.cell(row=rowoffset*3+1,column=1).value = 'mse'
    ws.cell(row=rowoffset*4+1,column=1).value = 'N'
    #ws.cell(row=rowoffset+1,column=1).font =  xl.styles.Font(italic=True)
    #ws.cell(row=rowoffset2+1,column=1).font = xl.styles.Font(italic=True)

    for i,solvent in enumerate(args.solvents):
        ws.cell(row=1,column=i*2+2).value = solvent.capitalize()
        if i == len(args.solvents)-1:
            lastcol = ws.cell(row=1,column=i*2+2).column

    for i,group in enumerate(supergroups,2):
        ws.cell(row=i,column=1).value = group
        for m in range(1,5):
            ws.cell(row=i+rowoffset*m,column=1).value = group

        for j,stats in enumerate(statslist[group]):
            if stats is None : continue
            ws.cell(row=i,column=j*2+2).value = stats.value
            ws.cell(row=i,column=j*2+3).value = stats.std
            ws.cell(row=i+rowoffset,column=j*2+2).value = stats.analytic
            ws.cell(row=i+rowoffset*2,column=j*2+2).value = stats.pvalue
            ws.cell(row=i+rowoffset*3,column=j*2+2).value = stats.mse
            ws.cell(row=i+rowoffset*4,column=j*2+2).value = stats.n
    """ws.conditional_formatting.add(
        "B%d:%s%d"%(2,lastcol,1+len(dblib.supergroups)),
        xl.formatting.CellIsRule(operator='greaterThan', formula=['0.5'],
        font=xl.styles.Font(bold=True)))
    ws.conditional_formatting.add(
        "B%d:%s%d"%(rowoffset*2+2,lastcol,rowoffset*2+1+len(dblib.supergroups)),
        xl.formatting.CellIsRule(operator='lessThan', formula=['0.05'],
        font=xl.styles.Font(bold=True)))"""

    wb.save(args.xls)
