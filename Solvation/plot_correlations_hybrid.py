# Author: Samuel Genheden samuel.genheden@gmail.com

"""
Program to plot the correlation between calculated and experimental free energies

Examples:
  plot_correlations.py -xls raw.xlsx -solvents hexane octane nonane hexanol octanol nonanol water
"""

import argparse
import os

import openpyxl as xl
import numpy as np
import matplotlib
import matplotlib.pylab as plt
import scipy.stats as stats

from sgenlib import colors

def _extract_nsolutes(ws):
    # G3 contains the MAD formula
    range_str = ws['G3'].value[9:-1]
    first,last = map(lambda s:int(s[1:]),range_str.split(':'))
    return last - first + 1

def _extract_data(ws):
    nsolutes = _extract_nsolutes(ws)
    return np.asarray([(ws.cell(row=i,column=2).value,
                        ws.cell(row=i,column=4).value)
                            for i in range(3,nsolutes+3)])

def _extract_reldata(ws):
    nsolutes = _extract_nsolutes(ws)
    return np.asarray([((ws.cell(row=i,column=12).value -
                                ws.cell(row=i,column=2).value)/(2.3*8.314/1000*298),
                            (ws.cell(row=i,column=14).value-
                                ws.cell(row=i,column=4).value)/(2.3*8.314/1000*298))
                        for i in range(3,nsolutes+3)])

def _extract_memdata(wb,offset=0):
    def getval(cellval):
        if isinstance(cellval,float):
            return cellval
        else:
            pos = cellval.index("!")
            return wb[cellval[1:pos]][cellval[pos+1:]].value
    ws = wb['mem']
    exp = np.asarray([ws.cell(row=i+offset,column=2).value for i in range(17,28)])
    umbrella = np.asarray([ws.cell(row=i+offset,column=3).value for i in range(17,28)])
    hexane = np.asarray([(getval(ws.cell(row=i+offset,column=6).value) -
                                getval(ws.cell(row=i+offset,column=2).value))/(2.3*8.314/1000*298)
                            for i in range(3,14)])
    octanol = np.asarray([(getval(ws.cell(row=i+offset,column=6).value) -
                                getval(ws.cell(row=i+offset,column=4).value))/(2.3*8.314/1000*298)
                            for i in range(3,14)])
    return [exp,umbrella,hexane,octanol]

def _plot_data(figure, datalist, labels, ylabels, xlabels, ncols=4, minv=None, maxv=None):
    if isinstance(ylabels,str):
        ylabels = [ylabels]*len(datalist)
        xlabels = [xlabels]*len(datalist)
        ilabels = False
    else:
        ilabels = True

    nrows = int(np.ceil(len(datalist)/float(ncols)))
    if minv is None :
        minv = np.floor(min([d.min() for d in datalist]))
    if maxv is None :
        maxv = np.ceil(max([d.max() for d in datalist]))
    print "Limits = %.2f %.2f"%(minv, maxv)
    vrange = maxv - minv
    delta = np.ceil(vrange/4.0)
    ticks = np.arange(minv,maxv+1.0,delta)

    for i,(data,label,ylabel,xlabel) in enumerate(zip(datalist,labels,ylabels,xlabels),1):
        a = figure.add_subplot(nrows,ncols,i)
        fit = np.polyfit(data[:,0],data[:,1],1)
        a.plot([fit[0]*minv+fit[1],fit[0]*maxv+fit[1]],[minv,maxv],"-k")
        a.plot([minv,maxv],[minv,maxv],"--k")    
        a.plot(data[:,1],data[:,0],".",color=colors.color(i-1))
        tau,tpval = stats.kendalltau(data[:,1],data[:,0])
        r,rpval = stats.pearsonr(data[:,1],data[:,0])
        print "%s\t%.5f\t%.5f\t%.5f\t%5f"%(label,r,rpval,tau,tpval)

        a.set_xlim(minv,maxv)
        a.set_ylim(minv,maxv)
        a.set_xticks(ticks)
        a.set_yticks(ticks)
        if ilabels or a.is_first_col():
            a.set_ylabel(ylabel)
        if ilabels or a.is_last_row():
            a.set_xlabel(xlabel)
        a.text(0.05,0.90,label,transform=a.transAxes, fontsize='9')
        a.set_aspect('equal')

    figure.tight_layout()
    print ""

if __name__ == '__main__':

    argparser = argparse.ArgumentParser(description="Plot correlations")
    argparser.add_argument('-xls','--xls',help="the filename of the XLS file")
    argparser.add_argument('-solvents','--solvents',nargs="+",help="the solvents in the XLS file to analyse")
    argparser.add_argument('-out','--out',help="the output filename")
    argparser.add_argument('-postfix','--postfix',help="the sheet name postfix")
    argparser.add_argument('-labels','--labels',nargs="+",help="special labels")
    argparser.add_argument('-limits','--limits',nargs=2,type=float,help="plot limits",default=[None,None])
    args = argparser.parse_args()

    matplotlib.rcParams.update({'font.size': 9})

    wb = None
    try :
        wb = xl.load_workbook(filename = args.xls)
    except :
        print "Could not open the XLS file. Exit"
        quit()

    sheets = []
    for solvent in args.solvents:
        try :
            postfix = args.postfix if solvent == 'water' else 'water'+args.postfix
            ws = wb['raw_'+solvent+postfix]
        except :
            raise Exception("Unable to find sheet for solvent %s"%('raw_'+solvent+postfix))
        else:
            sheets.append(ws)

    figure = plt.figure(1,figsize=(7,2.5))
    datalist = [_extract_data(sheet) for sheet in sheets]
    labels = list(args.solvents)
    if args.labels is not None :
        for special_label in args.labels :
            idx,lbl = special_label.split(":")
            labels[int(idx)-1] = lbl
    _plot_data(figure,datalist,labels,
        r'$\Delta G_\mathrm{calc.} \mathrm{[kJ/mol]}$',r'$\Delta G_\mathrm{exp.} \mathrm{[kJ/mol]}$',
        ncols=4, minv=args.limits[0], maxv=args.limits[1])
    figure.savefig(args.out,format="png",dpi=300)

    figure = plt.figure(2,figsize=(4.0,2.5))
    sheets2 = [sheet for sheet,solvent in zip(sheets,args.solvents) if solvent not in ["water", "waterpol"]]
    solvents2 = [solvent for solvent in args.solvents if solvent not in ["water", "waterpol"]]
    datalist = [_extract_reldata(sheet) for sheet in sheets2]
    _plot_data(figure,datalist,solvents2,
        r'$\mathrm{log} P_\mathrm{calc.}$',r'$\mathrm{log} P_\mathrm{exp.}$',ncols=2)
    figure.savefig(os.path.splitext(args.out)[0]+"_parti.png",format="png",dpi=300)
